import os
import csv
import json
import time
import tempfile
from typing import Dict, Any, List, Tuple, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


AUDIT_JSON_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "my_blog_engagement_audit.json"))

BUDDY_FIELDNAMES_MASTER = [
    "블로그ID",
    "닉네임",
    "블로그링크",
    "그룹",
    "이웃구분",
    "추가일",
    "내 새글보기",
    "반응한 글 수",
    "공감한 글 수",
    "댓글 단 글 수",
    "총댓글 수",
    "관찰 범위",
    "확인 시각",
    "반응 구분",
]

BUDDY_FIELDNAMES_REACTED = [
    "순위",
    "블로그ID",
    "닉네임",
    "블로그링크",
    "그룹",
    "이웃구분",
    "추가일",
    "반응한 글 수",
    "공감한 글 수",
    "댓글 단 글 수",
    "총댓글 수",
    "관찰 범위",
    "확인 시각",
]

BUDDY_FIELDNAMES_UNRESPONSIVE = [
    "블로그ID",
    "닉네임",
    "블로그링크",
    "그룹",
    "이웃구분",
    "추가일",
    "신규유예구분",
    "내 새글보기",
    "반응한 글 수",
    "관찰 범위",
    "확인 시각",
]


class EngagementAuditStore:
    """
    내 블로그 전체 이웃 기준 감사 및 반응자/무반응자 통합 리포트 저장소 (V13.1)
    - 1. 종합 엑셀 파일 (.xlsx): data/buddy_engagement_audit_YYYYMMDD.xlsx (전체이웃, 반응이웃_랭킹, 무반응이웃_관리 시트)
    - 2. Master 이웃 감사 CSV: data/buddy_engagement_audit_YYYYMMDD.csv
    - 3. 무반응 이웃 전용 CSV: data/unresponsive_buddies_YYYYMMDD.csv
    - 4. 종합 감사 JSON: data/my_blog_engagement_audit.json
    """

    @classmethod
    def get_file_paths(cls) -> Tuple[str, str, str, str]:
        date_str = time.strftime("%Y%m%d")
        data_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data"))
        os.makedirs(data_dir, exist_ok=True)

        json_path = os.path.join(data_dir, "my_blog_engagement_audit.json")
        excel_path = os.path.join(data_dir, f"buddy_engagement_audit_{date_str}.xlsx")
        master_csv = os.path.join(data_dir, f"buddy_engagement_audit_{date_str}.csv")
        unresp_csv = os.path.join(data_dir, f"unresponsive_buddies_{date_str}.csv")

        return json_path, excel_path, master_csv, unresp_csv

    @staticmethod
    def _safe_cell(value: Any) -> Any:
        """Keep spreadsheet formula-like external text literal."""
        if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
            return "'" + value
        return value

    @classmethod
    def _atomic_json(cls, path: str, payload: Dict[str, Any]) -> None:
        directory = os.path.dirname(path)
        fd, tmp = tempfile.mkstemp(prefix=".audit-", suffix=".json", dir=directory)
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=False, indent=2)
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(tmp, path)
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    @classmethod
    def _atomic_csv(cls, path: str, fieldnames: List[str], rows: List[Dict[str, Any]]) -> None:
        directory = os.path.dirname(path)
        fd, tmp = tempfile.mkstemp(prefix=".audit-", suffix=".csv", dir=directory)
        try:
            with os.fdopen(fd, "w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    writer.writerow({key: cls._safe_cell(row.get(key, "")) for key in fieldnames})
                handle.flush()
                os.fsync(handle.fileno())
            os.replace(tmp, path)
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    @classmethod
    def format_master_row(cls, row: Dict[str, Any]) -> Dict[str, Any]:
        b_id = row.get("blog_id", "")
        blog_url = row.get("blog_url", f"https://m.blog.naver.com/{b_id}") if b_id else ""
        return {
            "블로그ID": b_id,
            "닉네임": row.get("nickname", ""),
            "블로그링크": blog_url,
            "그룹": row.get("group_name") or "기본그룹",
            "이웃구분": row.get("buddy_type", "이웃"),
            "추가일": row.get("added_date", ""),
            "내 새글보기": row.get("news_feed_state", "확인 불가"),
            "반응한 글 수": row.get("engaged_post_count", 0),
            "공감한 글 수": row.get("like_count", 0),
            "댓글 단 글 수": row.get("comment_count", 0),
            "총댓글 수": row.get("comment_entry_count", 0),
            "관찰 범위": row.get("observation_scope", "최근 5개 글"),
            "확인 시각": row.get("checked_at", ""),
            "반응 구분": row.get("reaction_category", "확인 불가"),
        }

    @classmethod
    def format_reacted_row(cls, rank: int, row: Dict[str, Any]) -> Dict[str, Any]:
        b_id = row.get("blog_id", "")
        blog_url = row.get("blog_url", f"https://m.blog.naver.com/{b_id}") if b_id else ""
        return {
            "순위": rank,
            "블로그ID": b_id,
            "닉네임": row.get("nickname", ""),
            "블로그링크": blog_url,
            "그룹": row.get("group_name") or "기본그룹",
            "이웃구분": row.get("buddy_type", "이웃"),
            "추가일": row.get("added_date", ""),
            "반응한 글 수": row.get("engaged_post_count", 0),
            "공감한 글 수": row.get("like_count", 0),
            "댓글 단 글 수": row.get("comment_count", 0),
            "총댓글 수": row.get("comment_entry_count", 0),
            "관찰 범위": row.get("observation_scope", "최근 5개 글"),
            "확인 시각": row.get("checked_at", ""),
        }

    @classmethod
    def format_unresponsive_row(cls, row: Dict[str, Any]) -> Dict[str, Any]:
        b_id = row.get("blog_id", "")
        blog_url = row.get("blog_url", f"https://m.blog.naver.com/{b_id}") if b_id else ""
        grace_str = "신규 유예 (48시간 이내)" if row.get("is_recent_buddy") else "정리 대상 (무반응)"
        return {
            "블로그ID": b_id,
            "닉네임": row.get("nickname", ""),
            "블로그링크": blog_url,
            "그룹": row.get("group_name") or "기본그룹",
            "이웃구분": row.get("buddy_type", "이웃"),
            "추가일": row.get("added_date", ""),
            "신규유예구분": grace_str,
            "내 새글보기": row.get("news_feed_state", "확인 불가"),
            "반응한 글 수": row.get("engaged_post_count", 0),
            "관찰 범위": row.get("observation_scope", "최근 5개 글"),
            "확인 시각": row.get("checked_at", ""),
        }

    @classmethod
    def _export_excel_workbook(
        cls,
        excel_path: str,
        master_rows: List[Dict[str, Any]],
        reacted_rows: List[Dict[str, Any]],
        unresp_rows: List[Dict[str, Any]]
    ) -> None:
        """단일 엑셀 파일 내 3개 시트(전체종합, 반응자랭킹, 무반응관리) 생성 및 서식 적용"""
        if not OPENPYXL_AVAILABLE:
            return

        wb = openpyxl.Workbook()
        # 기본 시트 제거 후 이름 지정
        ws_master = wb.active
        ws_master.title = "전체이웃_종합"

        ws_reacted = wb.create_sheet(title="반응이웃_랭킹")
        ws_unresp = wb.create_sheet(title="무반응이웃_관리")

        # 스타일 정의
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        header_fill_blue = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_fill_green = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
        header_fill_red = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")

        data_font = Font(name="맑은 고딕", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0")
        )

        # 시트별 구성 헬퍼
        def populate_sheet(ws, fieldnames, data_list, header_fill):
            ws.views.sheetView[0].showGridLines = True
            # Header
            for col_idx, key in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_idx, value=key)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[1].height = 26

            # Data
            for row_idx, item in enumerate(data_list, 2):
                ws.row_dimensions[row_idx].height = 20
                for col_idx, key in enumerate(fieldnames, 1):
                    val = item.get(key, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=cls._safe_cell(val))
                    cell.font = data_font
                    cell.border = thin_border
                    if key in {"블로그ID", "닉네임", "블로그링크"}:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

            # Auto Column Widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or "")
                    max_len = max(max_len, len(val_str.encode('utf-8')) // 2 + len(val_str) // 2)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            # Freeze Panes
            ws.freeze_panes = "A2"

        # 1. 전체이웃_종합
        populate_sheet(ws_master, BUDDY_FIELDNAMES_MASTER, master_rows, header_fill_blue)
        # 2. 반응이웃_랭킹
        populate_sheet(ws_reacted, BUDDY_FIELDNAMES_REACTED, reacted_rows, header_fill_green)
        # 3. 무반응이웃_관리
        populate_sheet(ws_unresp, BUDDY_FIELDNAMES_UNRESPONSIVE, unresp_rows, header_fill_red)

        # 원자적 저장
        directory = os.path.dirname(excel_path)
        fd, tmp = tempfile.mkstemp(prefix=".audit-", suffix=".xlsx", dir=directory)
        os.close(fd)
        try:
            wb.save(tmp)
            os.replace(tmp, excel_path)
        finally:
            if os.path.exists(tmp):
                os.unlink(tmp)

    @classmethod
    def save_v13(cls, report: Dict[str, Any]) -> Tuple[str, str, str, str]:
        json_path, excel_path, master_csv, unresp_csv = cls.get_file_paths()

        # 1. 종합 JSON 저장
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        cls._atomic_json(json_path, report)

        # 2. Master Rows (전체 이웃)
        master_list = report.get("master_buddies", [])
        master_rows = [cls.format_master_row(b) for b in master_list]
        cls._atomic_csv(master_csv, BUDDY_FIELDNAMES_MASTER, master_rows)

        # 3. Reacted Rows (반응자 랭킹)
        reacted_buddies = [b for b in master_list if b.get("engaged_post_count", 0) > 0]
        reacted_buddies.sort(key=lambda x: (x.get("engaged_post_count", 0), x.get("comment_entry_count", 0), x.get("like_count", 0)), reverse=True)
        reacted_rows = [cls.format_reacted_row(idx, b) for idx, b in enumerate(reacted_buddies, 1)]

        # 4. Unresponsive Rows (무반응 이웃 전용)
        unresp_list = report.get("unresponsive_buddies", [])
        unresp_rows = [cls.format_unresponsive_row(u) for u in unresp_list]
        cls._atomic_csv(unresp_csv, BUDDY_FIELDNAMES_UNRESPONSIVE, unresp_rows)

        # 5. 다중 시트 Excel Workbook 통합 저장 (.xlsx)
        cls._export_excel_workbook(excel_path, master_rows, reacted_rows, unresp_rows)

        return json_path, excel_path, master_csv, unresp_csv

    @classmethod
    def save_v8(cls, report: Dict[str, Any]) -> Tuple[str, str, str, Optional[str]]:
        """Backward-compatible adapter for legacy callers."""
        json_path, excel_path, master_csv, unresp_csv = cls.save_v13(report)
        return json_path, master_csv, unresp_csv, None
